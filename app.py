from docx import Document
from flask import Flask, request
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
import re
from markupsafe import Markup


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


class Course(db.Model):
    id = db.Column(db.Integer(), primary_key=True)
    text = db.Column(db.String(900))


with app.app_context():
    db.create_all()


@app.route('/', methods=['POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        workbook = load_workbook(file)
        sheet = workbook.active

        for i in sheet.iter_rows(min_row=2, values_only=True):
            inp = Course(id=i[0], text=i[1])
            db.session.add(inp)
            db.session.commit()
    return "Hello"


@app.route('/docx', methods=['POST'])
def world():
    doc = request.files['file']
    doc_reader = Document(doc)
    data = ""
    for p in doc_reader.paragraphs:
        data += p.text + "\n"
    docx_input = Course(text=data) 
    db.session.add(docx_input)
    db.session.commit()
    return data



@app.route('/retrieve', methods=['GET'])
def retrieve_data():
    courses = Course.query.all()
    if courses:
        html_data = ''
        for course in courses:
            patterns = {
                'Email': r'Email:\s*([^\n]+)',
                'Mobile No': r'Mobile No:\s*([^\n]+)',
                'Location': r'Location:\s*([^\n]+)',
                'Score': r'Score:\s*([^\n]+)',
                'Gender': r'Gender\s*([^\n]+)',
                'Marital Status': r'Marital Status\s*([^\n]+)',
                'Date of birth': r'Date of birth\s*([^\n]+)',
                'Language': r'Language\s*([^\n]+)',
                'Objective': r'Objective\s*([^\n]+)',
                'Work Experience': r'Work Experience\s*([^\n]+)',
                'Academic Background': r'Academic Background\s*([^\n]+)',
                'Projects': r'Projects\s*([^\n]+)',
                'Technical Skills': r'Technical Skills\s*([^\n]+)',
                'Personal Details': r'Personal Details\s*([^\n]+)',
                'Links': r'Links\s*([^\n]+)',
                'Webside': r'Website:\s*([^\n]+)'
                # Add more patterns as needed
            }

            text = course.text

            for label, pattern in patterns.items():
                text = re.sub(pattern, f'<u><b>{label}:</b></u> \\1', text)
                

            html_data += f"{text}<br>"

        return Markup(html_data)
    else:
        return "No data found in the database"



if __name__ == "__main__":
    app.run(debug=True)
